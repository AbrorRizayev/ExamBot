
from django.contrib import admin
from django.core.exceptions import ValidationError
from django.db import transaction
from openpyxl import load_workbook

from .models import Exam, Question


@admin.register(Exam)
class ExamAdmin(admin.ModelAdmin):
    list_display = ('id', 'name', 'questions_count', 'created_at')
    search_fields = ('name',)

    def save_model(self, request, obj, form, change):
        with transaction.atomic():
            super().save_model(request, obj, form, change)
            obj.questions.all().delete()

            wb = load_workbook(obj.excel_file)
            ws = wb.active

            rows = list(ws.iter_rows(min_row=2, values_only=True))

            # if len(rows) != obj.questions_count:
            #     raise ValidationError(
            #         f"Excel savollar soni ({len(rows)}) "
            #         f"questions_count ({obj.questions_count}) ga teng emas"
            #     )

            questions = []

            for idx, row in enumerate(rows, start=2):
                question, a, b, c, d, correct = row

                if not all(row):
                    raise ValidationError(f"{idx}-qatorda bo‘sh katak bor")

                if correct not in ['A', 'B', 'C', 'D']:
                    raise ValidationError(
                        f"{idx}-qatorda noto‘g‘ri correct option: {correct}"
                    )

                questions.append(
                    Question(
                        exam=obj,
                        text=question,
                        option_a=a,
                        option_b=b,
                        option_c=c,
                        option_d=d,
                        correct_option=correct
                    )
                )

            Question.objects.bulk_create(questions)


